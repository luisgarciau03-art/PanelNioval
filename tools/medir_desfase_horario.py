"""Mide cuantas filas historicas pudieron guardarse con la fecha desplazada (M2).

Plan 5 · T5.3, punto 4. El contenedor corria en UTC y Mexico es UTC-6, asi que
lo capturado entre las 18:00 y las 23:59 hora local se guardaba con la fecha del
DIA SIGUIENTE y con una hora entre 00:00 y 05:59.

Este script SOLO CUENTA. No reescribe nada: corregir historico de clientes sin
que el owner lo pida es una operacion destructiva encubierta (SUPUESTO de §2.1
del plan, riesgo R7).

Lee el RESPALDO en disco, no las hojas de produccion: no hace ni una llamada de
red y no puede alterar nada por accidente.

Uso:  python tools/medir_desfase_horario.py docs/auditoria/respaldos/2026-09-04
"""
from __future__ import annotations

import collections
import pathlib
import re
import sys
from typing import Optional

import openpyxl

# La consola de Windows es cp1252 y revienta al imprimir lo que no existe ahi,
# llevandose por delante lo que viniera despues.
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

# Ventana sospechosa en hora local: lo escrito entre las 18:00 y medianoche de
# Mexico aparece guardado entre las 00:00 y las 06:00.
HORA_DESDE, HORA_HASTA = 0, 6

# Hojas que reciben un timestamp escrito por el panel, con la columna que lo
# lleva. Se identifican por encabezado, no por indice: reordenar la hoja
# rompe los proyectos externos sincronizados con ella, asi que el indice no es
# estable pero el nombre si.
OBJETIVOS = {
    "Respuestas de formulario 1": ["marca temporal", "fecha", "timestamp"],
    "ENVIOS_CATALOGO": ["fecha_solicitud", "timestamp_estado"],
    "PROSPECTOS BRUCE": ["fecha"],
}

# Hojas donde ADEMAS del panel escribe alguien mas. Solo esta: la columna A de
# 'Respuestas de formulario 1' la comparten Google Forms (que deja una celda de
# fecha real y NO sufre este bug) y el panel (que escribe TEXTO con
# value_input_option='RAW').
#
# La distincion por tipo de celda vale unicamente aqui. En ENVIOS_CATALOGO y
# PROSPECTOS BRUCE escriben solo el panel y el worker, asi que dar por "de otro
# origen" una celda de fecha en esas hojas la sacaria del numerador Y del
# denominador y produciria un 0 % que significa "no las conte", no "no hay
# riesgo". Si aparece una celda de fecha ahi (por ejemplo porque el respaldo
# paso por Excel y auto-tipifico el texto), se cuenta como del panel y se avisa.
HOJAS_CON_OTRO_ORIGEN = {"Respuestas de formulario 1"}

# dd/mm/aaaa hh:mm[:ss]  — el formato que escriben app.py y nucleo_catalogo.
PATRON = re.compile(r"^\s*(\d{1,2})/(\d{1,2})/(\d{4})[ ,]+(\d{1,2}):(\d{2})")


def hora_de(valor: object, hoja: str) -> tuple[Optional[int], bool, bool]:
    """Devuelve (hora 0-23, la_escribio_el_panel, tipo_inesperado).

    `tipo_inesperado` marca una celda de fecha en una hoja donde solo escriben
    el panel y el worker: se sigue contando como del panel, pero se reporta,
    porque significa que el respaldo no conserva el tipo original y la
    clasificacion por tipo dejo de ser fiable ahi.
    """
    if valor is None:
        return None, False, False

    es_celda_de_fecha = hasattr(valor, "hour") and hasattr(valor, "year")
    if es_celda_de_fecha:
        if hoja in HOJAS_CON_OTRO_ORIGEN:
            return valor.hour, False, False      # la puso Google Forms
        return valor.hour, True, True            # aqui solo escribe el panel

    m = PATRON.match(str(valor))
    if not m:
        return None, False, False
    return int(m.group(4)), True, False          # texto: lo escribio el panel


def medir(directorio: pathlib.Path) -> int:
    total = collections.Counter()
    hubo_columna = False
    avisos: list[str] = []

    for ruta in sorted(directorio.glob("*.xlsx")):
        try:
            wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        except Exception as e:                   # xlsx corrupto o que no lo es
            # Se avisa y se sigue con los demas: tumbar la corrida entera por un
            # archivo ilegible daria menos informacion, no mas.
            avisos.append(f"no se pudo abrir {ruta.name}: {type(e).__name__}: {e}")
            continue

        for hoja, columnas in OBJETIVOS.items():
            if hoja not in wb.sheetnames:
                continue
            ws = wb[hoja]
            filas = ws.iter_rows(values_only=True)
            try:
                encabezados = [str(c or "").strip().lower() for c in next(filas)]
            except StopIteration:
                continue

            indices = {c: encabezados.index(c) for c in columnas if c in encabezados}
            if not indices:
                continue
            hubo_columna = True

            cuenta = collections.Counter()
            for fila in filas:
                for col, i in indices.items():
                    if i >= len(fila):
                        continue
                    celda = fila[i]
                    if celda is None or (isinstance(celda, str) and not celda.strip()):
                        continue
                    cuenta[(col, "no_vacias")] += 1

                    h, del_panel, inesperado = hora_de(celda, hoja)
                    if h is None:
                        continue
                    cuenta[(col, "con_hora")] += 1
                    if inesperado:
                        cuenta[(col, "tipo_inesperado")] += 1
                    if del_panel:
                        cuenta[(col, "del_panel")] += 1
                        if HORA_DESDE <= h < HORA_HASTA:
                            cuenta[(col, "sospechosa")] += 1

            for col in indices:
                no_vacias = cuenta[(col, "no_vacias")]
                con_hora = cuenta[(col, "con_hora")]

                # "Columna encontrada pero nada legible" NO es lo mismo que
                # "columna vacia", y callarlo daria un cero de los que este
                # proyecto ya se ha comido varias veces.
                if no_vacias and not con_hora:
                    avisos.append(
                        f"{hoja} :: {col}: {no_vacias} celdas con dato y NINGUNA "
                        f"con hora reconocible. ¿Cambio el formato de fecha? "
                        f"El cero de esta columna no es un cero valido."
                    )
                if not con_hora:
                    continue

                panel = cuenta[(col, "del_panel")]
                sosp = cuenta[(col, "sospechosa")]
                inesperado = cuenta[(col, "tipo_inesperado")]
                pct = (100.0 * sosp / panel) if panel else 0.0

                print(f"  {hoja} :: {col}")
                print(f"      filas con hora            : {con_hora}")
                print(f"      escritas por el panel     : {panel}"
                      f"   (otro origen: {con_hora - panel})")
                print(f"      del panel, 00:00-05:59    : {sosp}  ({pct:.1f} %)")
                if inesperado:
                    print(f"      OJO: {inesperado} celdas de tipo fecha en una hoja "
                          f"donde solo escribe el panel (se cuentan como suyas)")

                total["con_hora"] += con_hora
                total["del_panel"] += panel
                total["sospechosa"] += sosp
        wb.close()

    if not hubo_columna:
        print("ERROR: no se encontro ninguna columna de timestamp conocida.")
        print("El barrido no puede devolver 0 valido si no alcanza el dato.")
        for a in avisos:
            print(f"  aviso: {a}")
        return 2

    panel = total["del_panel"]
    sosp = total["sospechosa"]
    pct = (100.0 * sosp / panel) if panel else 0.0
    print()
    print(f"TOTAL filas con hora        : {total['con_hora']}")
    print(f"TOTAL escritas por el panel : {panel}   <- el universo en riesgo")
    print(f"TOTAL sospechosas           : {sosp}  ({pct:.1f} % de las del panel)")

    if avisos:
        print()
        print("AVISOS (el resultado de arriba puede estar incompleto):")
        for a in avisos:
            print(f"  - {a}")

    print()
    print("Sospechosa = escrita por el panel entre 00:00 y 05:59, la ventana")
    print("donde caen las capturas de 18:00-23:59 hora de Mexico. NO es prueba")
    print("de desfase fila a fila: una captura real de madrugada cae igual ahi.")
    print()
    print("Y es una cota floja por arriba en otro sentido: solo las filas")
    print("escritas desde el contenedor (UTC) pudieron desplazarse. Las escritas")
    print("desde la PC del owner ya estaban en hora de Mexico, y el respaldo no")
    print("distingue que proceso escribio cada fila.")
    print()
    print("NO se corrige nada. La decision sobre el historico es del owner.")
    return 0


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print(__doc__)
        sys.exit(1)
    destino = pathlib.Path(sys.argv[1])
    if not destino.is_dir():
        print(f"ERROR: no existe el directorio {destino}")
        sys.exit(1)
    sys.exit(medir(destino))
